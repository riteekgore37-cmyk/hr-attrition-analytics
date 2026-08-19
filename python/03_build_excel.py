"""
Builds the Excel deliverable: HR_Attrition_Analysis.xlsx
  Sheet 1: Raw Data (employees table)
  Sheet 2: Summary (SUMIFS/COUNTIFS-driven KPIs by department)
  Sheet 3: Dashboard (charts referencing the Summary sheet)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

DATA = "../data/"
OUT = "../excel/HR_Attrition_Analysis.xlsx"

employees = pd.read_csv(DATA + "employees.csv")

wb = Workbook()

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5597")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ws1 = wb.active
ws1.title = "Raw Data"
cols = list(employees.columns)
for j, col in enumerate(cols, start=1):
    c = ws1.cell(row=1, column=j, value=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center")
for i, row in enumerate(employees.itertuples(index=False), start=2):
    for j, val in enumerate(row, start=1):
        cell = ws1.cell(row=i, column=j, value=val)
        cell.font = BODY_FONT
ws1.freeze_panes = "A2"
for j, col in enumerate(cols, start=1):
    width = max(12, min(22, employees[col].astype(str).str.len().max() + 2))
    ws1.column_dimensions[get_column_letter(j)].width = width

n = len(employees)

ws2 = wb.create_sheet("Summary")
ws2["A1"] = "HR Attrition — Department Summary"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:F1")

headers = ["Department", "Headcount", "Employees Left", "Attrition Rate %", "Avg Monthly Income", "Avg Tenure (Years)"]
for j, h in enumerate(headers, start=1):
    c = ws2.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

departments = sorted(employees["Department"].unique())
dept_col = cols.index("Department") + 1
attr_col = cols.index("Attrition") + 1
income_col = cols.index("MonthlyIncome") + 1
tenure_col = cols.index("TenureYears") + 1

dept_L = get_column_letter(dept_col)
attr_L = get_column_letter(attr_col)
income_L = get_column_letter(income_col)
tenure_L = get_column_letter(tenure_col)
rng = f"'Raw Data'!${dept_L}$2:${dept_L}${n+1}"
attr_rng = f"'Raw Data'!${attr_L}$2:${attr_L}${n+1}"
income_rng = f"'Raw Data'!${income_L}$2:${income_L}${n+1}"
tenure_rng = f"'Raw Data'!${tenure_L}$2:${tenure_L}${n+1}"

for i, dept in enumerate(departments, start=4):
    ws2.cell(row=i, column=1, value=dept).font = BODY_FONT
    ws2.cell(row=i, column=2, value=f'=COUNTIF({rng},A{i})').font = BODY_FONT
    ws2.cell(row=i, column=3, value=f'=COUNTIFS({rng},A{i},{attr_rng},"Yes")').font = BODY_FONT
    ws2.cell(row=i, column=4, value=f'=ROUND(C{i}/B{i}*100,1)').font = BODY_FONT
    ws2.cell(row=i, column=5, value=f'=ROUND(SUMIFS({income_rng},{rng},A{i})/B{i},0)').font = BODY_FONT
    ws2.cell(row=i, column=6, value=f'=ROUND(SUMIFS({tenure_rng},{rng},A{i})/B{i},1)').font = BODY_FONT
    for col in range(1, 7):
        ws2.cell(row=i, column=col).border = BORDER

last_row = 3 + len(departments)
totrow = last_row + 1
ws2.cell(row=totrow, column=1, value="Company-wide").font = Font(name="Arial", bold=True)
ws2.cell(row=totrow, column=2, value=f"=SUM(B4:B{last_row})").font = Font(name="Arial", bold=True)
ws2.cell(row=totrow, column=3, value=f"=SUM(C4:C{last_row})").font = Font(name="Arial", bold=True)
ws2.cell(row=totrow, column=4, value=f"=ROUND(C{totrow}/B{totrow}*100,1)").font = Font(name="Arial", bold=True)
ws2.cell(row=totrow, column=5, value=f"=ROUND(AVERAGE(E4:E{last_row}),0)").font = Font(name="Arial", bold=True)
ws2.cell(row=totrow, column=6, value=f"=ROUND(AVERAGE(F4:F{last_row}),1)").font = Font(name="Arial", bold=True)

for j in range(1, 7):
    ws2.column_dimensions[get_column_letter(j)].width = 20

ws3 = wb.create_sheet("Dashboard")
ws3["A1"] = "HR Attrition Dashboard"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:H1")
ws3["A2"] = "Charts below are linked to the Summary sheet — update Raw Data and everything recalculates."
ws3["A2"].font = Font(name="Arial", italic=True, size=9, color="808080")

bar = BarChart()
bar.title = "Attrition Rate by Department (%)"
bar.y_axis.title = "Attrition %"
bar.x_axis.title = "Department"
data_ref = Reference(ws2, min_col=4, min_row=3, max_row=last_row)
cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=last_row)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
bar.height = 9
bar.width = 18
ws3.add_chart(bar, "A4")

pie = PieChart()
pie.title = "Headcount by Department"
data_ref2 = Reference(ws2, min_col=2, min_row=3, max_row=last_row)
pie.add_data(data_ref2, titles_from_data=True)
pie.set_categories(cats_ref)
pie.height = 9
pie.width = 12
ws3.add_chart(pie, "A22")

bar2 = BarChart()
bar2.title = "Average Monthly Income by Department"
bar2.y_axis.title = "Avg Monthly Income (INR)"
data_ref3 = Reference(ws2, min_col=5, min_row=3, max_row=last_row)
bar2.add_data(data_ref3, titles_from_data=True)
bar2.set_categories(cats_ref)
bar2.height = 9
bar2.width = 18
ws3.add_chart(bar2, "K4")

wb.save(OUT)
print("Saved:", OUT)